"""Combine working supplementary CSV files into one XLSX workbook."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


TABLES = [
    ("S1_Datasets", "Table_S1_dataset_inventory_working.csv", "Dataset inventory and evidence role"),
    ("S2_Profiles", "Table_S2_anisonet_pinn_profile_working.csv", "anisoNET PINN profiles and intended use"),
    ("S3_Metrics", "Table_S3_metric_definitions_working.csv", "Metric definitions and caveats"),
    ("S4A_Apoe_QC", "Table_S4A_gse193107_Apoe_batch_metrics_working.csv", "GSE193107 Apoe field-QC summary"),
    ("S4B_Gfap_QC", "Table_S4B_gse193107_Gfap_batch_metrics_working.csv", "GSE193107 Gfap field-QC summary"),
    ("S4C_Targets", "Table_S4C_gse193107_target_batch_comparison_working.csv", "Apoe/Gfap target comparison"),
    ("S5_Synthetic", "Table_S5_synthetic_barrier_metrics_working.csv", "Synthetic barrier benchmark metrics"),
    ("S6_CrossTissue", "Table_S6_cross_tissue_evidence_summary_working.csv", "Cross-tissue evidence summary"),
    ("S7A_KidneyBoundary", "Table_S7A_kidney_evidence_boundary_working.csv", "Kidney evidence-boundary diagnostics"),
    ("S7B_KidneyScreen", "Table_S7B_kidney_marker_screen_working.csv", "Kidney marker-screen follow-up"),
    ("S8A_RuntimeDefault", "Table_S8A_resource_profile_default_working.csv", "Runtime profile for default runs"),
    ("S8B_RuntimeLowPDE", "Table_S8B_resource_profile_low_pde_working.csv", "Runtime profile for low-PDE runs"),
]


def add_readme_sheet(workbook: Workbook, table_dir: Path) -> None:
    ws = workbook.create_sheet("README", 0)
    rows = [
        ["Supplementary Tables Working Workbook"],
        ["Generated for the anisoNET GPB revision."],
        ["Table directory", str(table_dir)],
        ["Interpretation note", "Fitted-source metrics should not be described as held-out prediction performance."],
        ["Interpretation note", "Cross-tissue tables define portability and claim boundaries, not universal superiority."],
        [],
        ["Sheet", "Source CSV", "Title / role"],
    ]
    for row in rows:
        ws.append(row)
    for sheet_name, filename, title in TABLES:
        ws.append([sheet_name, filename, title])

    ws["A1"].font = Font(bold=True, size=14)
    for cell in ws[7]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
    ws.freeze_panes = "A8"
    widths = [26, 62, 70]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def add_csv_sheet(workbook: Workbook, sheet_name: str, csv_path: Path) -> None:
    ws = workbook.create_sheet(sheet_name)
    if not csv_path.exists():
        ws.append(["MISSING_SOURCE", str(csv_path)])
        return

    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        for row in reader:
            ws.append(row)

    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, ws.max_column + 1):
        max_len = 8
        for row_idx in range(1, min(ws.max_row, 200) + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_len = max(max_len, min(len(str(value)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--table-dir", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)

    add_readme_sheet(workbook, args.table_dir)

    for sheet_name, filename, _title in TABLES:
        add_csv_sheet(workbook, sheet_name, args.table_dir / filename)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
